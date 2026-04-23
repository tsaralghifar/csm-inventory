#!/usr/bin/env python3
"""
Generate Excel report untuk Permintaan Material (openpyxl)
Usage: python export_pm_excel.py <json_file> <output_path>
"""

import sys, json, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Helpers ───────────────────────────────────────────────────────────────────
def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip('#'))

def bdr(style="thin", color="C8D0DA"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def bdr_med(color="1A3A5C"):
    s = Side(style="medium", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def no_bdr():
    return Border()

def fmt_date(val):
    if not val: return '-'
    try:
        d = datetime.fromisoformat(str(val)[:10])
        months = ['Januari','Februari','Maret','April','Mei','Juni',
                  'Juli','Agustus','September','Oktober','November','Desember']
        return f"{d.day:02d} {months[d.month-1]} {d.year}"
    except:
        return str(val)

def set_cell(ws, row, col, value='', font=None, fill=None, align=None, border=None, height=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    if height: ws.row_dimensions[row].height = height
    return c

# ── Preset styles ─────────────────────────────────────────────────────────────
NAVY  = "1A3A5C"
BLUE  = "2563A8"
LBLUE = "E8EDF4"
LGRAY = "F5F7FA"
STRIP = "F1F5F9"
WHITE = "FFFFFF"
GRAY  = "64748B"
DARK  = "1F2937"
AMBER = "FFFBEB"
AMB_B = "F59E0B"

fH1   = Font(name='Arial', bold=True, size=15, color=WHITE)
fH2   = Font(name='Arial', bold=True, size=10, color=WHITE)
fSec  = Font(name='Arial', bold=True, size=9,  color=NAVY)
fLbl  = Font(name='Arial', size=9,  color=GRAY)
fVal  = Font(name='Arial', bold=True, size=9, color=NAVY)
fVal2 = Font(name='Arial', size=9,  color=DARK)
fTH   = Font(name='Arial', bold=True, size=9, color=WHITE)
fTD   = Font(name='Arial', size=9,  color=DARK)
fTDB  = Font(name='Arial', bold=True, size=9, color=DARK)
fMono = Font(name='Courier New', bold=True, size=9, color=NAVY)
fSign = Font(name='Arial', bold=True, size=9, color=NAVY)
fNote = Font(name='Arial', size=9, color='92400E')

aL  = Alignment(horizontal='left',   vertical='center', indent=1)
aLC = Alignment(horizontal='left',   vertical='center')
aR  = Alignment(horizontal='right',  vertical='center')
aC  = Alignment(horizontal='center', vertical='center')
aCB = Alignment(horizontal='center', vertical='bottom')

def generate_pm_excel(pm, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Permintaan Material"

    # ── Lebar kolom (12 kolom) ────────────────────────────────────────────────
    # Layout info: [lbl:A+B] [val:C+D+E] [gap:F] [lbl:G+H] [val:I+J+K+L]
    # Layout tabel: A(#) B(PartNo) C+D(Nama) E(KodeUnit) F(TipeUnit) G(Qty) H(Satuan) I+J+K+L(Ket)
    col_w = [4, 14, 18, 14, 10, 4, 14, 12, 10, 10, 8, 16]
    #         A   B    C   D   E  F   G   H   I   J  K   L
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    N = 12  # total kolom

    is_part = pm.get('type') == 'part'
    stmap = {
        'draft':'Draft','pending_chief':'Menunggu Chief Mekanik',
        'pending_manager':'Menunggu Manager','pending_ho':'Menunggu Admin HO',
        'manager_approved':'Disetujui Manager','approved':'Disetujui HO',
        'purchasing':'Proses Purchasing','completed':'Selesai','rejected':'Ditolak',
    }
    status = stmap.get(pm.get('status',''), pm.get('status',''))
    doctype = "MATERIAL REQUEST PART" if is_part else "MATERIAL REQUEST OFFICE"

    r = 1

    # ── Row 1: Header perusahaan ──────────────────────────────────────────────
    ws.row_dimensions[r].height = 30
    ws.merge_cells(f'A{r}:L{r}')
    set_cell(ws,r,1,"PT. CIPTA SARANA MAKMUR", fH1, hfill(NAVY), aL, bdr_med(NAVY))
    r += 1

    # ── Row 2: Judul ──────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'A{r}:L{r}')
    set_cell(ws,r,1,f"{doctype}  ·  {pm.get('nomor','')}  ·  {status.upper()}", fH2, hfill(BLUE), aL)
    r += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 6
    r += 1

    # ── Section header INFO & PERSETUJUAN ─────────────────────────────────────
    ws.row_dimensions[r].height = 18
    # INFO: A+B+C+D+E (5 col)
    ws.merge_cells(f'A{r}:E{r}')
    set_cell(ws,r,1,"INFORMASI PERMINTAAN", fSec, hfill(LBLUE), aL, bdr_med(NAVY))
    for col in range(2,6): ws.cell(r,col).border = bdr_med(NAVY); ws.cell(r,col).fill = hfill(LBLUE)
    # GAP: F
    ws.cell(r,6).border = no_bdr(); ws.cell(r,6).fill = hfill(WHITE)
    # PERSETUJUAN: G+H+I+J+K+L (6 col)
    ws.merge_cells(f'G{r}:L{r}')
    set_cell(ws,r,7,"PERSETUJUAN", fSec, hfill(LBLUE), aL, bdr_med(NAVY))
    for col in range(8,13): ws.cell(r,col).border = bdr_med(NAVY); ws.cell(r,col).fill = hfill(LBLUE)
    r += 1

    # ── Info rows ─────────────────────────────────────────────────────────────
    wh      = (pm.get('warehouse') or {}).get('name','-')
    req     = (pm.get('requester') or {}).get('name','-')
    chief   = (pm.get('chiefAuthorizer') or {}).get('name','-')
    manager = (pm.get('managerApprover') or {}).get('name','-')
    ho      = (pm.get('approver') or pm.get('hoApprover') or {}).get('name','-')
    needed  = fmt_date(pm.get('needed_date'))
    created = fmt_date(pm.get('created_at'))

    info = [
        ("No. PM",         pm.get('nomor','-'), "Tgl. Dibutuhkan", needed),
        ("Gudang / Site",  wh,                  "Chief Mekanik",   chief),
        ("Diajukan Oleh",  req,                 "Manager",         manager),
        ("Tanggal Dibuat", created,              "Admin HO",        ho),
    ]

    for lbl1, val1, lbl2, val2 in info:
        ws.row_dimensions[r].height = 16
        # Label kiri: A+B merged
        ws.merge_cells(f'A{r}:B{r}')
        set_cell(ws,r,1, lbl1, fLbl, hfill(LGRAY), aL, bdr())
        ws.cell(r,2).border = bdr(); ws.cell(r,2).fill = hfill(LGRAY)
        # Value kiri: C+D+E merged
        ws.merge_cells(f'C{r}:E{r}')
        set_cell(ws,r,3, val1, fVal, hfill(WHITE), aL, bdr())
        ws.cell(r,4).border = bdr(); ws.cell(r,5).border = bdr()
        # Gap F
        ws.cell(r,6).border = no_bdr(); ws.cell(r,6).fill = hfill(WHITE)
        # Label kanan: G+H merged
        ws.merge_cells(f'G{r}:H{r}')
        set_cell(ws,r,7, lbl2, fLbl, hfill(LGRAY), aL, bdr())
        ws.cell(r,8).border = bdr(); ws.cell(r,8).fill = hfill(LGRAY)
        # Value kanan: I+J+K+L merged
        ws.merge_cells(f'I{r}:L{r}')
        set_cell(ws,r,9, val2, fVal2, hfill(WHITE), aL, bdr())
        for col in [10,11,12]: ws.cell(r,col).border = bdr()
        r += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 6
    r += 1

    # ── Header tabel barang ───────────────────────────────────────────────────
    # Kolom tabel: A(#) B(PartNo) C+D(Nama) E(KodeUnit) F(TipeUnit) G(Qty) H(Satuan) I+J+K+L(Ket)
    ws.row_dimensions[r].height = 20
    tbl_headers = [
        (1,  1,  '#',                      aC),
        (2,  2,  'Part Number',             aC),
        (3,  4,  'Nama Barang / Deskripsi', aL),
        (5,  5,  'Kode Unit',               aC),
        (6,  6,  'Tipe Unit',               aC),
        (7,  7,  'Qty',                     aC),
        (8,  8,  'Satuan',                  aC),
        (9,  12, 'Keterangan',              aL),
    ]
    for (c1, c2, label, align) in tbl_headers:
        if c1 != c2:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=label)
        cell.font      = fTH
        cell.fill      = hfill(NAVY)
        cell.alignment = align
        cell.border    = bdr(color="0F2440")
        if c1 != c2:
            for col in range(c1+1, c2+1):
                ws.cell(r,col).border = bdr(color="0F2440")
                ws.cell(r,col).fill   = hfill(NAVY)
    r += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    items = pm.get('items', [])
    for idx, item in enumerate(items):
        ws.row_dimensions[r].height = 16
        bg = STRIP if idx % 2 == 1 else WHITE

        row_data = [
            (1,  1,  idx+1,                                                               aC,  fTD,   False),
            (2,  2,  item.get('part_number') or (item.get('item') or {}).get('part_number') or '-', aC, fMono, False),
            (3,  4,  item.get('nama_barang') or '-',                                      aL,  fTDB,  False),
            (5,  5,  item.get('kode_unit') or '-',                                        aC,  fTD,   False),
            (6,  6,  item.get('tipe_unit') or '-',                                        aC,  fTD,   False),
            (7,  7,  item.get('qty'),                                                     aC,  fTD,   False),
            (8,  8,  item.get('satuan') or '-',                                           aC,  fTD,   False),
            (9,  12, item.get('keterangan') or '-',                                       aL,  fTD,   False),
        ]
        for (c1, c2, val, align, font, _) in row_data:
            if c1 != c2:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            cell = ws.cell(row=r, column=c1, value=val)
            cell.font      = font
            cell.fill      = hfill(bg)
            cell.alignment = align
            cell.border    = bdr()
            if c1 != c2:
                for col in range(c1+1, c2+1):
                    ws.cell(r,col).fill   = hfill(bg)
                    ws.cell(r,col).border = bdr()
        r += 1

    # ── Catatan ───────────────────────────────────────────────────────────────
    notes = pm.get('notes')
    if notes:
        ws.row_dimensions[r].height = 6; r += 1
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f'A{r}:L{r}')
        cell = ws.cell(row=r, column=1, value=f"Catatan: {notes}")
        cell.font      = fNote
        cell.fill      = hfill(AMBER)
        cell.alignment = aL
        cell.border    = Border(
            left   = Side(style='thick',  color=AMB_B),
            right  = Side(style='thin',   color='C8D0DA'),
            top    = Side(style='thin',   color='C8D0DA'),
            bottom = Side(style='thin',   color='C8D0DA'),
        )
        r += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 12
    r += 1

    # ── Tanda tangan 4 kolom ─────────────────────────────────────────────────
    # Layout: [A+B+C sign1] [gap D] [E+F+G sign2] [gap H] [I+J sign3] [K+L sign4]
    chief_name    = (pm.get('chiefAuthorizer') or {}).get('name', '')
    approved_name = ((pm.get('approver') or {}).get('name') or
                     (pm.get('managerApprover') or {}).get('name') or '')

    signs = [
        (1,  3,  'ORDERED BY LOGISTIC',    ''),
        (5,  7,  'RECEIVED BY PURCHASING', ''),
        (9,  10, 'AUTHORIZED BY',          chief_name),
        (11, 12, 'APPROVED BY',            approved_name),
    ]
    gaps_h = [4, 8]  # kolom gap

    # Header tanda tangan
    ws.row_dimensions[r].height = 18
    for (c1, c2, label, _) in signs:
        if c1 != c2:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=label)
        cell.font = fSign; cell.fill = hfill(LBLUE)
        cell.alignment = aC; cell.border = bdr_med(NAVY)
        if c1 != c2:
            for col in range(c1+1, c2+1):
                ws.cell(r,col).fill = hfill(LBLUE)
                ws.cell(r,col).border = bdr_med(NAVY)
    for gc in gaps_h:
        ws.cell(r,gc).border = no_bdr()
        ws.cell(r,gc).fill   = hfill(WHITE)
    r += 1

    # Box tanda tangan (tinggi 55)
    ws.row_dimensions[r].height = 55
    for (c1, c2, _, val) in signs:
        if c1 != c2:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=val)
        cell.font = fSign; cell.fill = hfill(WHITE)
        cell.alignment = aCB; cell.border = bdr_med(NAVY)
        if c1 != c2:
            for col in range(c1+1, c2+1):
                ws.cell(r,col).fill   = hfill(WHITE)
                ws.cell(r,col).border = bdr_med(NAVY)
    for gc in gaps_h:
        ws.cell(r,gc).border = no_bdr()
        ws.cell(r,gc).fill   = hfill(WHITE)

    # ── Print setup ───────────────────────────────────────────────────────────
    ws.page_setup.orientation  = 'landscape'
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_margins.left       = 0.5
    ws.page_margins.right      = 0.5
    ws.page_margins.top        = 0.75
    ws.page_margins.bottom     = 0.75

    wb.save(out_path)
    print(f"OK:{out_path}")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: export_pm_excel.py <json_file> <output_path>")
        sys.exit(1)
    arg1 = sys.argv[1]
    if os.path.isfile(arg1):
        with open(arg1, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
    else:
        data = json.loads(arg1)
    generate_pm_excel(data, sys.argv[2])